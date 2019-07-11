from itertools import combinations
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
from matplotlib import style
import pandas as pd
import numpy as np
from FuncAni_2 import RST_Parameter_Calc
import time
from datetime import datetime
import openpyxl

base_name = "RSTlog"
now = datetime.now()    # current date and time
suffix = now.strftime("%d-%m-%Y_%H-%M-%S")
file_name = "_".join([base_name, suffix]) + ".xlsx"
wb = openpyxl.Workbook()   # Workbook() creates a new excel file
ws = wb.active  # fetches the currently active worksheet
headings = ['Timestamp', 'Decision Value', 'n(LA)',  'n(UA)', 'Accuracy', 'Stability Index']
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 24
row = 1
for col, entry in enumerate(headings, start=1): # filling in the headings
        ws.cell(row=row, column=col, value=entry)
        wb.save(file_name)


# accepting file_path
flag = False
while not flag:
    file_path = input("Please enter the complete correct file path of the dataset: \n")
    ch = (input("Press 'Y' to confirm \nPress 'N' to enter file path again \n")).upper()
    if ch == 'Y':
        flag = True
    elif ch == 'N':
        flag = False

style.use('fivethirtyeight')
fig = plt.figure()
ax1 = fig.add_subplot(1,1,1)
try:
    def animate(i):

        global row, file_name
    
        # df = pd.read_csv(file_path, index_col=0, delimiter = ' ')    # reading csv file data as a data frame variable
        df = pd.read_csv(file_path)    # reading csv file data as a data frame variable
        # print(df)
        # print()

        index = df.index
        columns = df.columns
        values = df.values
        num_col = df.shape[1]
        num_row = df.shape[0]


        obj_elem_set = set()    # stores serial number of each object as set elements
        for i in range(1, df.shape[0]+1):
            obj_elem_set.add(i)
        # print(obj_elem_set)
        # print()
    
        # storage variables used
        dict_col = {}   # stores cardinal nos. of element values column-wise as well as unique data-wise
        elemen_list = []    # stores elementary list numbers of conditional attributes
        cris_list = []  # stores crisp set numbers of decision attribute
        list_col = []   # stores names of columns
        list_combi = []     # stores combinations of columns as tuples
        elem_indiscern_2_list = []  # stores List of all indiscernible combinations taking double conditional attributes
        elem_list = []  # stores elementary set numbers of conditional attributes
        dict_low = {}   # stores lower bound set
        dict_upp = {}   # stores upper bound set
        dict_accu = {}  # stores accuracy of parameters
        dict_SI = {}    # stores stability index of parameters
        dict_boun = {}  # stores boundary region of each combinations
        dict_out = {}   # stores outside region of each combinations
        elem_dict = {}  # stores elementary list and dict numbers of conditional attributes
        elemen_dict = {}    # stores elementary dict numbers of conditional attributes
        dict_indiscern_2 = {}   # stores elementary list for double conditional attributes
        dict_nla = {}   # stores cardinal no. of Lower Approximation set
        dict_nua = {}   # stores cardinal no. of Upper Approximation set
    
        for column in columns:  # stores names of columns
            list_col.append(column)
        list_col.pop()  # don't include decision attribute (last column)

        len_combi = num_col - 1
        col_combi = combinations(list_col, len_combi)   # stores combinations of columns taken 'len_combi' at a time
        list_combi = list(col_combi)    # stores combinations of columns as tuples
        # print(list_combi)


        # PandasAgeWalkFunc class object created
        obj_item = RST_Parameter_Calc(df)

        # obtain complete serial numbers of all unique conditional and decision attributes as a dictionary
        dict_col = obj_item.col_item_split()
        # print(dict_col)

        # obtain elementary set and crisp set
        elem_dict = obj_item.elem_list(dict_col)
        elemen_list = elem_dict['Elem List']
        cris_list = elemen_list.pop()
        elemen_dict = elem_dict['Elem Dict']
        rem_key = columns[-1]
        elemen_dict.pop(rem_key)
    

        # print("Elementary List for Single-Conditional Attributes: \n" + str(elemen_list) + "\n")
        # print("Crisp List: " + str(cris_list) + "\n")
        # print("Elementary Dictionary: " + str(elemen_dict) + "\n")

        # Returns elementary list for multiple conditional attributes
        dict_indiscern_2 = obj_item.column_combinations(elemen_dict, list_combi)
        # print("List of all indiscernible combinations taking multiple conditional attributes is as follows: ")
        # print(str(dict_indiscern_2) + "\n")

        for val in dict_indiscern_2.values():
            elem_indiscern_2_list.append(val)
        # print("Elementary List for Multi-Conditional Attributes: \n" + str(elem_indiscern_2_list) + "\n")
        elem_list = elem_indiscern_2_list   # for multiple conditional attributes

        dec_items = sorted(list(set(df[columns[-1]].unique())))
        len_dec = len(dec_items)

        # print("Lower and Upper Approximations are given below: ")
        for i in range(0, len_dec): # calculating the RST Parameters
            dec_val = dec_items[i]
            # print("RST LA, UA for Decision Attribute Value: " + str(dec_val))
            dict_low = obj_item.low_approx(dec_val, dict_col, elem_list, list_combi)    # obtain lower approximation
            dict_upp = obj_item.upp_approx(dec_val, dict_col, elem_list, list_combi)    # obtain upper approximation
            # print("Lower Approximation: " + str(dict_low) + "\n")
            # print("Upper Approximation: " + str(dict_upp) + "\n")
            for key in dict_low.keys():  # length of either dict_low or dict_upp; both are equal to no. of CAs taken at a time
                dict_nla[dec_val] = len(dict_low[key])    # no. of elements in la
                dict_nua[dec_val] = len(dict_upp[key])    # no. of elements in ua
            dict_accu[dec_val] = obj_item.get_accu(dict_low, dict_upp)  # obtain accuracy parameter using accuracy = nLa/nUa
            dict_SI[dec_val] = obj_item.get_SI(dict_low, dict_upp, len(obj_elem_set))  # obtain stability index(SI) parameter using SI = (n_la + n_ua + 1)/(n + 1) - 0.5
            dict_boun[dec_val] = obj_item.get_boundary(dict_low, dict_upp)  # get boundary region
            dict_out[dec_val] = obj_item.get_outside_region(obj_elem_set, dict_upp) # get outside region

        print("Accuracy of the parameters for each decision attribute is given below: ")
        print(str(dict_accu) + "\n")
        print("Stability Index(SI) of the parameters for each decision attribute is given below: ")
        print(str(dict_SI) + "\n")
        # print("Boundary region is: Upper Approx. - Lower Approx. = ")
        # print(str(dict_boun) + "\n")
        # print("Outside region is: Universal Set - Upper Approx: ")
        # print(str(dict_out) + "\n")
    
        new_row = []
        wb = openpyxl.load_workbook(filename=file_name)
        ws = wb.active
    
        for i in range(0, len_dec):
            now = datetime.now()    # current date and time
            date_time = now.strftime("%d/%m/%Y, %H:%M:%S")
            print(date_time)
            print()
            dec_val = dec_items[i]
            row = row + 1
            new_row = [date_time, dec_val, dict_nla[dec_val], dict_nua[dec_val], dict_accu[dec_val], dict_SI[dec_val]]
            for col, entry in enumerate(new_row, start=1):
                ws.cell(row=row, column=col, value=entry)
                wb.save(file_name)
            
    ani = FuncAnimation(fig, animate, interval=5000)
    plt.show()

except PermissionError:
    pass
